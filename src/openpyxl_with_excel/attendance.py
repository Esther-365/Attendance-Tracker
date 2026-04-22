import os
from dotenv import load_dotenv
import openpyxl
import smtplib
from email.message import EmailMessage
from pathlib import Path
import re
import time

load_dotenv(Path(__file__).parent.parent.parent / ".env")
DEBUG_MODE = os.getenv("DEBUG_MODE", "False").lower() == "true"
host = os.getenv("HOST")
#.getenv returns a string, so we need to parse it to int cause the smtplib.SMTP expects port as int.
port = int(os.getenv("PORT"))
username = os.getenv("MAIL_USERNAME")
password = os.getenv("MAIL_PASSWORD")
pattern = re.compile(r"^([a-z]+_[a-z]+)") 

wb = openpyxl.load_workbook(os.getenv("ATTENDANCE_PATH"))

def send_email(server, stud_email, info, prof_email=None):
    #Info[Student ID, Level,Number of Leaves, Course name,Staff name if prof_email is not None]
    stud_name = pattern.search(stud_email).group(0).replace("_", " ").title()
    
    try:
        if info[2] == 2:
            msg = EmailMessage()
            msg['subject'] = f"Attendance Warning - {info[3]}" 
            msg['From'] = "emmacol@eu.edu"
            msg['To'] = stud_email
            body = f"Dear {stud_name},\n\n\tThis letter serves as an official warning regarding your attendance for {info[3]}. You have currently accumulated 2 leaves, which is a cause for concern. Please note, that exceeding 3 leaves will result in you being barred from writing the final exam for this course. We urge you take this seriously and make necessary adjustments going forward.\n\nStudent ID: {info[0]} Level: {info[1]}\n\nBest regards,\nEmmanuel's University"
            msg.set_content(body)
            server.send_message(msg)

        elif info[2] == 3:
            recepients =  [stud_email, prof_email]
            for rep in recepients:
                msg = EmailMessage()
                msg['subject'] = f"Attendance Warning - {info[3]}" 
                msg['From'] = "emmacol@eu.edu"
                msg['To'] = rep
                if rep == stud_email:
                    body = f"Dear {stud_name},\n\n\tThis letter serves as an official warning regarding your attendance for {info[3]}. You have currently accumulated 3 leaves, which is a cause for concern. Please note, that exceeding 3 leaves will result in you being barred from writing the final exam for this course. We urge you take this seriously and make necessary adjustments going forward.\n\nStudent ID: {info[0]} Level: {info[1]}\n\nBest regards,\nEmmanuel's University"
                else:
                    body = f"Dear {info[4]},\n\n\tThis letter serves as an official warning regarding the attendance of your student, {stud_name} for {info[3]}. They have currently accumulated 3 leaves, which is a cause for concern. Please note, that exceeding 3 leaves will result in them being barred from writing the final exam for this course. We urge you to reach out to the student and encourage them to make necessary adjustments going forward.\n\nStudent ID: {info[0]} Level: {info[1]} Student Email: {stud_email}\n\nBest regards,\nEmmanuel's University"
                msg.set_content(body)
                server.send_message(msg)
                time.sleep(12)

        else:
            recepients = [stud_email, prof_email]
            for rep in recepients:
                msg = EmailMessage()
                msg['subject'] = f"Attendance Warning - {info[3]}" 
                msg['From'] = "emmacol@eu.edu"
                msg['To'] = rep
                if rep == stud_email:
                    body = f"Dear {stud_name},\n\n\tThis letter serves as an official warning regarding your attendance for {info[3]}. You have currently accumulated {info[2]} leaves. Kindly note, that haven exceeding 3 leaves resulted in you being barred from writing the final exam for this course. We urge you reach out to your academic advisor and make necessary adjustments going forward.\n\nStudent ID: {info[0]} Level: {info[1]}\n\nBest regards,\nEmmanuel's University"
                else:
                    body = f"Dear {info[4]},\n\n\tThis letter serves as an official warning regarding the attendance of your student, {stud_name} for {info[3]}. They have currently accumulated {info[2]} leaves, which exceeds the minimum requirement. Please note, that having exceeded 3 leaves will result in them being barred from writing the final exam for this course. We urge you to reach out to the student and encourage them to make necessary adjustments going forward.\n\nStudent ID: {info[0]} Level: {info[1]} Student Email: {stud_email}\n\nBest regards,\nEmmanuel's University"
                msg.set_content(body)
                server.send_message(msg)
                time.sleep(12)
        
    except Exception as e:
        print(f"An error occurred while sending the email: {e}")

with smtplib.SMTP(host,port) as server:
    server.starttls()
    server.login(username, password)

    for sheet in wb.worksheets:
        #note it's 1-indexed
        for cell in sheet[1]:
            if cell.value is not None and "Leaves" in cell.value:
                col_val = cell.column #This gives the column number of the cell
                for row_index,rowd in enumerate(sheet.iter_rows(values_only=True, min_row = 2, min_col = col_val, max_col = col_val), start=2):
                    #rowd returns a tuple of values in the loop range
                    core = rowd[0]
                    #core gets the first item in rowd cause it's only one item based on how I wrote the loop.
                    #row_index is the index of the row that core is in.

                    #Info[Student ID, Level,Number of Leaves, Course name,Staff name if prof_email is not None]
                    if core == 2:
                        stud_email = sheet.cell(row = row_index,column= 3).value
                        info = [sheet.cell(row = row_index, column = 1).value, sheet.cell(row = row_index, column = 2).value, 2,cell.value[:-6].strip()]
                        time.sleep(12) #This is to avoid sending too many emails in a short period of time, which can cause the email server to block our account.        
                        send_email(server,stud_email, info) 
                    
                    elif core == 3:
                        stud_email = sheet.cell(row = row_index,column= 3).value
                        prof_email = sheet.cell(row = row_index, column = col_val+2).value
                        info = [sheet.cell(row = row_index, column = 1).value, sheet.cell(row = row_index, column = 2).value, 3,cell.value[:-6], sheet.cell(row = row_index, column = col_val+1).value]
                        time.sleep(12) #This is to avoid sending too many emails in a short period of time, which can cause the email server to block our account.
                        send_email(server,stud_email, info, prof_email)
                    
                    elif core >= 4:
                        stud_email = sheet.cell(row = row_index,column= 3).value
                        prof_email = sheet.cell(row = row_index, column = col_val+2).value
                        info = [sheet.cell(row = row_index, column = 1).value, sheet.cell(row = row_index, column = 2).value, core,cell.value[:-6], sheet.cell(row = row_index, column = col_val+1).value]
                        time.sleep(12) #This is to avoid sending too many emails in a short period of time, which can cause the email server to block our account.
                        send_email(server,stud_email, info, prof_email)




